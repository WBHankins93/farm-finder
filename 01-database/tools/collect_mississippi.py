#!/usr/bin/env python3
"""Collect Mississippi direct-to-consumer farm candidates with three source passes.

The collector never edits the canonical workbook. It writes a staged, source-backed
candidate release that must be reviewed before promotion.
"""

from __future__ import annotations

import csv
import difflib
import hashlib
import html
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CANONICAL = ROOT / "research" / "local_farm_database_final.xlsx"
OUTPUT_DIR = ROOT / "research" / "ms-expansion"
USER_AGENT = "FarmFinder/1.0 (+quarterly public-directory verification)"
TODAY = date.today().isoformat()

GENUINE_ARCHIVES = {
    "Genuine MS — Grown": "https://genuinems.com/members/grown/",
    "Genuine MS — Raised": "https://genuinems.com/members/raised/",
}
MDAC_VENDOR_LIST = "https://agnet.mdac.ms.gov/Website/vendorlist"
MDAC_MARKETPLACE = "https://agnet.mdac.ms.gov/MarketPortal/MarketPortal"
PICK_YOUR_OWN = {
    "PickYourOwn — North": "https://www.pickyourown.org/MSnorth.htm",
    "PickYourOwn — Jackson/West Central": "https://www.pickyourown.org/MSjackson.htm",
    "PickYourOwn — East Central": "https://www.pickyourown.org/MSeast.htm",
    "PickYourOwn — Southeast": "https://www.pickyourown.org/MSse.htm",
    "PickYourOwn — Southwest": "https://www.pickyourown.org/MSsw.htm",
}


@dataclass
class Candidate:
    farm_name: str
    state: str = "MS"
    city: str = ""
    county: str = ""
    county_source: str = ""
    address: str = ""
    postal_code: str = ""
    products: str = ""
    phone: str = ""
    email: str = ""
    website_url: str = ""
    facebook_url: str = ""
    instagram_url: str = ""
    source_names: list[str] = field(default_factory=list)
    source_urls: list[str] = field(default_factory=list)
    source_modified_date: str = ""
    retrieved_date: str = TODAY
    source_passes: list[int] = field(default_factory=list)
    current_release_match: str = ""
    review_status: str = "Needs review"
    notes: str = ""


def clean_text(value: str) -> str:
    value = re.sub(r"<br\s*/?>", " | ", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    return re.sub(r"\s+", " ", html.unescape(value)).strip(" |\t\r\n")


def normalized_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()


def alias_name(value: str) -> str:
    tokens = normalized_name(value).split()
    drop = {
        "llc", "llp", "inc", "company", "co", "the", "and", "family",
        "farm", "farms", "ranch", "cattle", "livestock", "market",
    }
    return " ".join(token for token in tokens if token not in drop)


def fetch(url: str, attempts: int = 3, timeout: int = 35) -> tuple[str, dict]:
    errors: list[str] = []
    for attempt in range(1, attempts + 1):
        started = time.monotonic()
        try:
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=timeout) as response:
                body = response.read().decode(response.headers.get_content_charset() or "utf-8", "replace")
                return body, {
                    "url": url,
                    "attempts_used": attempt,
                    "http_status": response.status,
                    "bytes": len(body.encode("utf-8")),
                    "sha256": hashlib.sha256(body.encode("utf-8")).hexdigest(),
                    "elapsed_seconds": round(time.monotonic() - started, 3),
                    "error": "",
                }
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            errors.append(f"attempt {attempt}: {exc}")
            if attempt < attempts:
                time.sleep(0.8 * attempt)
    return "", {
        "url": url,
        "attempts_used": attempts,
        "http_status": 0,
        "bytes": 0,
        "sha256": "",
        "elapsed_seconds": 0,
        "error": " | ".join(errors),
    }


def append_source(candidate: Candidate, name: str, url: str, source_pass: int) -> None:
    if name not in candidate.source_names:
        candidate.source_names.append(name)
    if url not in candidate.source_urls:
        candidate.source_urls.append(url)
    if source_pass not in candidate.source_passes:
        candidate.source_passes.append(source_pass)


def merge_candidate(target: Candidate, incoming: Candidate) -> None:
    for field_name in (
        "city", "county", "address", "postal_code", "products", "phone",
        "email", "website_url", "facebook_url", "instagram_url",
        "source_modified_date", "notes",
    ):
        if not getattr(target, field_name) and getattr(incoming, field_name):
            setattr(target, field_name, getattr(incoming, field_name))
    for source_name, source_url in zip(incoming.source_names, incoming.source_urls):
        if source_name not in target.source_names:
            target.source_names.append(source_name)
        if source_url not in target.source_urls:
            target.source_urls.append(source_url)
    target.source_passes = sorted(set(target.source_passes + incoming.source_passes))


def parse_genuine_archive(body: str, source_name: str, source_url: str) -> list[Candidate]:
    candidates: list[Candidate] = []
    for article in re.findall(r"<article\b.*?</article>", body, flags=re.I | re.S):
        match = re.search(
            r"<h5>(.*?)</h5>.*?<a\s+href=[\"']([^\"']*/directory/[^\"']+)[\"'][^>]*>\s*<h3>(.*?)</h3>",
            article,
            flags=re.I | re.S,
        )
        if not match:
            continue
        location, profile_url, farm_name = map(clean_text, match.groups())
        city = re.sub(r",?\s*(MS|Mississippi)\s*$", "", location, flags=re.I).strip(" ,")
        candidate = Candidate(farm_name=farm_name, city=city)
        append_source(candidate, source_name, profile_url, 1)
        candidates.append(candidate)
    return candidates


def parse_genuine_profile(candidate: Candidate, body: str) -> Candidate:
    contact_match = re.search(r'<span\s+class="member-contact">(.*?)</span>', body, flags=re.I | re.S)
    contact = clean_text(contact_match.group(1)) if contact_match else ""
    if contact:
        parts = [part.strip() for part in contact.split("|") if part.strip()]
        if parts:
            candidate.address = parts[0]
        if len(parts) > 1:
            candidate.phone = parts[-1]
        zip_match = re.search(r"\b(\d{5})(?:-\d{4})?\b", contact)
        if zip_match:
            candidate.postal_code = zip_match.group(1)
        city_match = re.search(r",\s*([^,|]+),\s*(?:MS|Mississippi)\s+\d{5}", contact, flags=re.I)
        if city_match:
            candidate.city = city_match.group(1).strip()

    link_classes = {
        "website_url": "prov-web",
        "email": "prov-email",
        "facebook_url": "prov-facebook",
        "instagram_url": "prov-instagram",
    }
    for field_name, css_class in link_classes.items():
        match = re.search(
            rf'<li\s+class="{css_class}".*?<a\s+href="([^"]+)"',
            body,
            flags=re.I | re.S,
        )
        if match:
            value = html.unescape(match.group(1)).strip()
            if field_name == "email":
                value = re.sub(r"^mailto:", "", value, flags=re.I).split("?", 1)[0]
            setattr(candidate, field_name, value)

    products = [clean_text(item) for item in re.findall(r'<li\s+class="product-item">(.*?)</li>', body, flags=re.I | re.S)]
    candidate.products = "; ".join(item for item in products if item)
    modified = re.search(r'<meta\s+property="article:modified_time"\s+content="([^"]+)"', body, flags=re.I)
    if modified:
        candidate.source_modified_date = modified.group(1)[:10]
    return candidate


def parse_mdac_vendors(body: str) -> list[Candidate]:
    candidates: list[Candidate] = []
    row_pattern = re.compile(
        r'<tr[^>]*>\s*<td[^>]*>.*?<a\s+href="([^"]*FarmerMarket_VendorView[^\"]*)"[^>]*>.*?<font[^>]*>(.*?)</font>.*?</a>.*?</td>\s*<td[^>]*>(.*?)</td>',
        flags=re.I | re.S,
    )
    for href, name, product_type in row_pattern.findall(body):
        farm_name = clean_text(name)
        product_type = clean_text(product_type)
        agricultural = product_type in {"Fruit and Vegetables", "Dairy", "Meat", "Nursery"}
        agricultural = agricultural or bool(re.search(r"farm|garden|honey|fung|grow|meadow", farm_name, flags=re.I))
        if not agricultural:
            continue
        detail_url = urllib.parse.urljoin(MDAC_VENDOR_LIST, href)
        candidate = Candidate(farm_name=farm_name, products=product_type)
        append_source(candidate, "MDAC Mississippi Farmers Market vendors", detail_url, 2)
        candidates.append(candidate)
    return candidates


def parse_mdac_vendor_detail(candidate: Candidate, body: str) -> Candidate:
    ids = {
        "address": "MainContent_Label_Address",
        "email": "MainContent_Label_Email",
        "phone": "MainContent_Label_BusPhone",
        "products": "MainContent_Label_AgProducts",
    }
    for field_name, element_id in ids.items():
        match = re.search(rf'<span\s+id="{element_id}">(.*?)</span>', body, flags=re.I | re.S)
        if match:
            value = clean_text(match.group(1))
            if value:
                setattr(candidate, field_name, value)
    return candidate


def parse_mdac_marketplace(body: str) -> list[Candidate]:
    candidates: list[Candidate] = []
    blocks = re.findall(r'<p\s+style="width:99%;text-align:justify">(.*?)</p>', body, flags=re.I | re.S)
    for block in blocks:
        match = re.search(
            r'<a\s+href="/MarketPortal/MarketPortal\?farm=[^"]+">(.*?)</a>\s+in\s+([^<(]+)\s*\(<mark[^>]*>(.*?)</mark>\s+county\)',
            block,
            flags=re.I | re.S,
        )
        if not match:
            continue
        farm_name, city, county = map(clean_text, match.groups())
        product_matches = [clean_text(value) for value in re.findall(r'>\s*Providing\s+([^<\r\n]+)</a>', block, flags=re.I)]
        other_products = [clean_text(value) for value in re.findall(r'MarketPortal_Single\?id=\d+">\s*([^<]+)</a>', block, flags=re.I)]
        phones = re.findall(r'href="tel:([^"]+)"', block, flags=re.I)
        emails = re.findall(r'href="mailto:\s*([^"]+)"', block, flags=re.I)
        external = [
            html.unescape(url).strip() for url in re.findall(r'<a\s+target="_blank"\s+href="([^"]+)"', block, flags=re.I)
        ]
        candidate = Candidate(
            farm_name=farm_name,
            city=city,
            county=county,
            county_source="MDAC listing",
            products="; ".join(dict.fromkeys(product_matches + other_products)),
            phone=phones[0].strip() if phones else "",
            email=emails[0].strip() if emails else "",
        )
        if external:
            url = external[0]
            if "facebook.com" in url:
                candidate.facebook_url = url
            else:
                candidate.website_url = url if re.match(r"https?://", url) else f"https://{url}"
        append_source(candidate, "MDAC Mississippi Farm Marketplace", MDAC_MARKETPLACE, 2)
        candidates.append(candidate)
    return candidates


def parse_pick_your_own(body: str, source_name: str, source_url: str) -> list[Candidate]:
    candidates: list[Candidate] = []
    county = ""
    token_pattern = re.compile(r'(<h3[^>]*>.*?</h3>|<li[^>]*>.*?</li>)', flags=re.I | re.S)
    for token in token_pattern.findall(body):
        if token.lower().startswith("<h3"):
            heading = clean_text(token)
            county = re.sub(r"\s+County$", "", heading, flags=re.I)
            continue
        text = clean_text(token)
        lowered = text.casefold()
        if "permanently closed" in lowered or "assumed permanently closed" in lowered:
            continue
        if not county or len(text) < 20:
            continue
        name_match = re.search(
            r'(?:class="farm"[^>]*>\s*(?:<a[^>]*>)?|<li[^>]*>\s*<b>\s*(?:<a[^>]*>)?)([^<]+)',
            token,
            flags=re.I | re.S,
        )
        if not name_match:
            continue
        farm_name = clean_text(name_match.group(1)).strip(" -")
        if not farm_name or farm_name.casefold().startswith(("click here", "please write")):
            continue
        city_match = re.search(r',\s*([^,]+),\s*MS\s+\d{5}', text, flags=re.I)
        phone_match = re.search(r'Phone:\s*([^.;]+)', text, flags=re.I)
        email_match = re.search(r'Email:\s*([^\s;]+@[^\s;]+)', text, flags=re.I)
        product_match = re.search(r'^.*?\s+-\s+(.*?)(?:\s+\d{2,5}\s|\s+Phone:|\s+Open:)', text)
        candidate = Candidate(
            farm_name=farm_name,
            county=county,
            county_source="PickYourOwn listing",
            city=city_match.group(1).strip() if city_match else "",
            products=product_match.group(1).strip(" ,") if product_match else "",
            phone=phone_match.group(1).strip() if phone_match else "",
            email=email_match.group(1).strip(" .") if email_match else "",
            notes="Conservative import: listing was not marked closed; confirm activity before promotion.",
        )
        append_source(candidate, source_name, source_url, 3)
        candidates.append(candidate)
    return candidates


def census_county(address: str) -> tuple[str, str]:
    if not address or not re.search(r"\bMS\b|Mississippi", address, flags=re.I):
        return "", ""
    query = urllib.parse.urlencode({
        "address": address,
        "benchmark": "Public_AR_Current",
        "vintage": "Current_Current",
        "format": "json",
    })
    body, log = fetch(f"https://geocoding.geo.census.gov/geocoder/geographies/onelineaddress?{query}")
    if not body:
        return "", log.get("error", "")
    try:
        payload = json.loads(body)
        matches = payload["result"]["addressMatches"]
        if not matches:
            return "", "no address match"
        county = matches[0]["geographies"]["Counties"][0]["BASENAME"]
        return str(county).strip(), ""
    except (KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        return "", f"unparsed Census response: {exc}"


def read_existing_release() -> tuple[dict[str, str], dict[str, str], list[dict[str, str]], int, int]:
    workbook = load_workbook(CANONICAL, read_only=True, data_only=True)
    sheet = workbook["All Farms"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    positions = {header: index for index, header in enumerate(headers)}
    exact: dict[str, str] = {}
    aliases: dict[str, str] = {}
    existing_records: list[dict[str, str]] = []
    source_rows = 0
    ms_rows = 0
    for row in rows:
        source_rows += 1
        name = str(row[positions["Farm Name"]] or "").strip()
        state = str(row[positions["State"]] or "").strip()
        if state == "MS":
            ms_rows += 1
            exact[normalized_name(name)] = name
            alias = alias_name(name)
            if alias:
                aliases[alias] = name
            existing_records.append({
                "name": name,
                "city": str(row[positions["City/Town"]] or "").strip(),
            })
    return exact, aliases, existing_records, source_rows, ms_rows


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source_log: list[dict] = []
    collected: list[Candidate] = []

    # Pass 1: statewide Genuine MS producer directories and current profiles.
    genuine_candidates: list[Candidate] = []
    for source_name, source_url in GENUINE_ARCHIVES.items():
        body, log = fetch(source_url)
        log.update({"pass": 1, "source_name": source_name, "retrieved_at": datetime.now(timezone.utc).isoformat()})
        parsed = parse_genuine_archive(body, source_name, source_url) if body else []
        log["records_parsed"] = len(parsed)
        source_log.append(log)
        genuine_candidates.extend(parsed)

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch, candidate.source_urls[0]): candidate for candidate in genuine_candidates}
        for future in as_completed(futures):
            candidate = futures[future]
            body, log = future.result()
            if body:
                parse_genuine_profile(candidate, body)
            log.update({
                "pass": 1,
                "source_name": "Genuine MS profile",
                "records_parsed": 1 if body else 0,
                "retrieved_at": datetime.now(timezone.utc).isoformat(),
            })
            source_log.append(log)
            collected.append(candidate)

    # Pass 2: MDAC market vendors and active commodity listings.
    vendor_body, vendor_log = fetch(MDAC_VENDOR_LIST)
    mdac_vendors = parse_mdac_vendors(vendor_body) if vendor_body else []
    vendor_log.update({
        "pass": 2,
        "source_name": "MDAC Mississippi Farmers Market vendors",
        "records_parsed": len(mdac_vendors),
        "retrieved_at": datetime.now(timezone.utc).isoformat(),
    })
    source_log.append(vendor_log)
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch, candidate.source_urls[0]): candidate for candidate in mdac_vendors}
        for future in as_completed(futures):
            candidate = futures[future]
            body, log = future.result()
            if body:
                parse_mdac_vendor_detail(candidate, body)
            log.update({
                "pass": 2,
                "source_name": "MDAC vendor profile",
                "records_parsed": 1 if body else 0,
                "retrieved_at": datetime.now(timezone.utc).isoformat(),
            })
            source_log.append(log)
            collected.append(candidate)

    marketplace_body, marketplace_log = fetch(MDAC_MARKETPLACE)
    marketplace_candidates = parse_mdac_marketplace(marketplace_body) if marketplace_body else []
    marketplace_log.update({
        "pass": 2,
        "source_name": "MDAC Mississippi Farm Marketplace",
        "records_parsed": len(marketplace_candidates),
        "retrieved_at": datetime.now(timezone.utc).isoformat(),
    })
    source_log.append(marketplace_log)
    collected.extend(marketplace_candidates)

    # Pass 3: all five Mississippi PickYourOwn regions, excluding rows marked closed.
    for source_name, source_url in PICK_YOUR_OWN.items():
        body, log = fetch(source_url)
        parsed = parse_pick_your_own(body, source_name, source_url) if body else []
        log.update({
            "pass": 3,
            "source_name": source_name,
            "records_parsed": len(parsed),
            "retrieved_at": datetime.now(timezone.utc).isoformat(),
        })
        source_log.append(log)
        collected.extend(parsed)

    # Consolidate exact duplicate names across passes; do not silently merge fuzzy names.
    consolidated: dict[str, Candidate] = {}
    for candidate in collected:
        key = normalized_name(candidate.farm_name)
        if not key:
            continue
        if key in consolidated:
            merge_candidate(consolidated[key], candidate)
        else:
            consolidated[key] = candidate

    # County-only enrichment from public source addresses. Exact coordinates are not stored.
    geocode_errors: list[dict[str, str]] = []
    geocode_targets = [candidate for candidate in consolidated.values() if not candidate.county and candidate.address]
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(census_county, candidate.address): candidate for candidate in geocode_targets}
        for future in as_completed(futures):
            candidate = futures[future]
            county, error = future.result()
            if county:
                candidate.county = county
                candidate.county_source = "U.S. Census geocoder from public source address"
            elif error:
                geocode_errors.append({"farm_name": candidate.farm_name, "address": candidate.address, "error": error})

    exact, aliases, existing_records, source_rows, existing_ms_rows = read_existing_release()
    for key, candidate in consolidated.items():
        if key in exact:
            candidate.current_release_match = exact[key]
            candidate.review_status = "Existing release row — enrich/verify"
        else:
            alias = alias_name(candidate.farm_name)
            if alias and alias in aliases:
                candidate.current_release_match = aliases[alias]
                candidate.review_status = "Possible existing alias — identity review"
            elif len(candidate.source_passes) >= 2:
                candidate.review_status = "New candidate — multi-source review"
            elif candidate.source_modified_date >= "2025-01-01":
                candidate.review_status = "New candidate — recent authoritative profile"
            else:
                best = None
                for existing in existing_records:
                    ratio = difflib.SequenceMatcher(None, normalized_name(candidate.farm_name), normalized_name(existing["name"])).ratio()
                    same_city = bool(candidate.city and existing["city"] and normalized_name(candidate.city) == normalized_name(existing["city"]))
                    score = ratio + (0.2 if same_city else 0)
                    if best is None or score > best[0]:
                        best = (score, ratio, same_city, existing["name"])
                if best and (best[1] >= 0.92 or (best[2] and best[1] >= 0.60)):
                    candidate.current_release_match = best[3]
                    candidate.review_status = "Possible existing fuzzy match — identity review"
                else:
                    candidate.review_status = "New candidate — needs independent verification"

    candidates = sorted(consolidated.values(), key=lambda item: normalized_name(item.farm_name))
    records = []
    for candidate in candidates:
        record = asdict(candidate)
        for field_name in ("source_names", "source_urls", "source_passes"):
            record[field_name] = " | ".join(map(str, record[field_name]))
        records.append(record)

    json_payload = {
        "release_id": f"ms-expansion-{TODAY}",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "canonical_workbook": str(CANONICAL),
        "canonical_source_rows": source_rows,
        "canonical_ms_rows": existing_ms_rows,
        "scope": "Publicly discoverable Mississippi farms/producers with direct-to-consumer, farm-market, agritourism, or branded producer evidence; not every USDA-defined farm.",
        "candidate_count": len(records),
        "records": records,
    }
    (OUTPUT_DIR / "mississippi-candidates.json").write_text(json.dumps(json_payload, indent=2), encoding="utf-8")
    with (OUTPUT_DIR / "mississippi-candidates.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]) if records else [])
        if records:
            writer.writeheader()
            writer.writerows(records)
    (OUTPUT_DIR / "source-pass-log.json").write_text(json.dumps(source_log, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "geocode-errors.json").write_text(json.dumps(geocode_errors, indent=2), encoding="utf-8")

    summary = {
        "status": "passed" if all(not entry["error"] for entry in source_log if entry["source_name"] in list(GENUINE_ARCHIVES) + list(PICK_YOUR_OWN) + ["MDAC Mississippi Farmers Market vendors", "MDAC Mississippi Farm Marketplace"]) else "completed_with_source_errors",
        "canonical_source_rows": source_rows,
        "canonical_ms_rows": existing_ms_rows,
        "three_passes_completed": sorted({entry["pass"] for entry in source_log}),
        "unique_candidates_collected": len(records),
        "existing_exact_or_alias_matches": sum(bool(record["current_release_match"]) for record in records),
        "new_candidates": sum(not bool(record["current_release_match"]) for record in records),
        "multi_source_candidates": sum(len(str(record["source_passes"]).split(" | ")) >= 2 for record in records),
        "source_requests": len(source_log),
        "failed_source_requests": sum(bool(entry["error"]) for entry in source_log),
        "county_values_present": sum(bool(record["county"]) for record in records),
        "census_geocode_failures": len(geocode_errors),
    }
    (OUTPUT_DIR / "collection-summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
