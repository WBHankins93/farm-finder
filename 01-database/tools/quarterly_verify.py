#!/usr/bin/env python3
"""Quarterly, exception-driven verification for FarmFinder releases and staging data.

This scanner retries every HTTP check up to three times. It writes a dated audit
report and never changes canonical farm values automatically.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import socket
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
MANIFEST = ROOT / "03-app" / "site" / "config" / "source-of-truth.json"
MS_CANDIDATES = ROOT / "research" / "ms-expansion" / "mississippi-candidates.json"
AUDIT_ROOT = ROOT / "research" / "quarterly-audits"
USER_AGENT = "FarmFinder/1.0 (+quarterly public-directory verification)"
TODAY = date.today().isoformat()


def normalized(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalized_name(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalized(value).casefold()).strip()


def with_scheme(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    if value.startswith(("http://", "https://")):
        return value
    return f"https://{value}"


def check_url(url: str, attempts: int = 3, timeout: int = 20) -> dict:
    errors: list[str] = []
    for attempt in range(1, attempts + 1):
        started = time.monotonic()
        try:
            request = urllib.request.Request(
                url,
                headers={"User-Agent": USER_AGENT, "Range": "bytes=0-8191"},
            )
            with urllib.request.urlopen(request, timeout=timeout) as response:
                body = response.read(8192)
                status = int(response.status)
                return {
                    "url": url,
                    "status": "reachable" if 200 <= status < 400 else "http_error",
                    "http_status": status,
                    "final_url": response.geturl(),
                    "attempts_used": attempt,
                    "elapsed_seconds": round(time.monotonic() - started, 3),
                    "sample_sha256": hashlib.sha256(body).hexdigest(),
                    "error": "",
                }
        except urllib.error.HTTPError as exc:
            # 401/403/405/429 still prove the host/resource route is reachable.
            if exc.code in {401, 403, 405, 429}:
                return {
                    "url": url,
                    "status": "reachable_restricted",
                    "http_status": exc.code,
                    "final_url": exc.geturl(),
                    "attempts_used": attempt,
                    "elapsed_seconds": round(time.monotonic() - started, 3),
                    "sample_sha256": "",
                    "error": str(exc),
                }
            errors.append(f"attempt {attempt}: HTTP {exc.code}")
            if exc.code in {404, 410}:
                break
        except (urllib.error.URLError, TimeoutError, socket.timeout, OSError) as exc:
            errors.append(f"attempt {attempt}: {exc}")
        if attempt < attempts:
            time.sleep(1.0 * attempt)
    return {
        "url": url,
        "status": "broken" if any("HTTP 404" in item or "HTTP 410" in item for item in errors) else "unreachable_after_3_attempts",
        "http_status": 404 if any("HTTP 404" in item for item in errors) else 410 if any("HTTP 410" in item for item in errors) else 0,
        "final_url": "",
        "attempts_used": min(attempts, len(errors) or attempts),
        "elapsed_seconds": 0,
        "sample_sha256": "",
        "error": " | ".join(errors),
    }


def canonical_audit() -> tuple[dict, list[dict], dict[str, set[str]]]:
    manifest = json.loads(MANIFEST.read_text())
    release = manifest["release"]
    workbook_path = (MANIFEST.parent.parent / release["workspacePath"]).resolve()
    digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[release["sheet"]]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalized(value) for value in next(rows)]
    positions = {header: index for index, header in enumerate(headers)}
    records = [dict(zip(headers, row)) for row in rows]
    errors: list[str] = []
    warnings: list[str] = []
    if digest != release["sha256"]:
        errors.append("canonical workbook checksum differs from manifest")
    if len(records) != release["sourceRowCount"]:
        errors.append(f"source row mismatch: expected {release['sourceRowCount']}, found {len(records)}")
    missing_columns = sorted(set(manifest["requiredColumns"]) - set(headers))
    if missing_columns:
        errors.append(f"missing required columns: {missing_columns}")

    missing_by_field: dict[str, int] = {}
    for field in manifest["requiredNonEmptyColumns"]:
        count = sum(not normalized(record.get(field)) or normalized(record.get(field)).casefold() in {"nan", "n/a", "unknown"} for record in records)
        missing_by_field[field] = count
        if count:
            errors.append(f"{field} has {count} missing values")

    names: dict[str, list[str]] = {}
    for record in records:
        names.setdefault(normalized_name(record.get("Farm Name")), []).append(normalized(record.get("Farm Name")))
    duplicate_groups = {key: values for key, values in names.items() if len(values) > 1}
    if duplicate_groups:
        warnings.append(f"{len(duplicate_groups)} normalized-name groups need identity review")

    url_owners: dict[str, set[str]] = {}
    row_findings: list[dict] = []
    for index, record in enumerate(records, start=2):
        website = with_scheme(normalized(record.get("Website URL")))
        if website:
            url_owners.setdefault(website, set()).add(f"canonical:{index}:{normalized(record.get('Farm Name'))}")
        findings = []
        if normalized(record.get("Parish/County")).casefold() in {"varies", "multiple"}:
            findings.append("county_not_specific")
        if not normalized(record.get("Contact Info")):
            findings.append("contact_missing")
        if normalized(record.get("Has Website")).casefold() == "yes" and not website:
            findings.append("website_flag_without_url")
        if findings:
            row_findings.append({
                "dataset": "canonical",
                "row": index,
                "farm_name": normalized(record.get("Farm Name")),
                "state": normalized(record.get("State")),
                "findings": " | ".join(findings),
            })

    report = {
        "workbook": str(workbook_path),
        "release_id": release["id"],
        "sha256": digest,
        "source_rows": len(records),
        "state_counts": {
            state: sum(normalized(record.get("State")) == state for record in records)
            for state in sorted({normalized(record.get("State")) for record in records})
        },
        "missing_required_by_field": missing_by_field,
        "duplicate_groups": duplicate_groups,
        "errors": errors,
        "warnings": warnings,
    }
    return report, row_findings, url_owners


def staging_audit(url_owners: dict[str, set[str]]) -> tuple[dict, list[dict]]:
    if not MS_CANDIDATES.exists():
        return {"candidate_file": str(MS_CANDIDATES), "status": "not_found"}, []
    payload = json.loads(MS_CANDIDATES.read_text())
    records = payload["records"]
    findings: list[dict] = []
    for index, record in enumerate(records, start=1):
        row_findings: list[str] = []
        for field in ("farm_name", "products", "source_urls", "retrieved_date"):
            if not normalized(record.get(field)):
                row_findings.append(f"missing_{field}")
        if not normalized(record.get("city")):
            row_findings.append("missing_city")
        if not normalized(record.get("county")):
            row_findings.append("missing_county")
        if normalized(record.get("review_status")).startswith("New candidate"):
            row_findings.append("not_promoted")
        if row_findings:
            findings.append({
                "dataset": "ms_staging",
                "row": index,
                "farm_name": normalized(record.get("farm_name")),
                "state": "MS",
                "findings": " | ".join(row_findings),
            })
        raw_urls = []
        for field in ("source_urls", "website_url", "facebook_url", "instagram_url"):
            raw_urls.extend(part.strip() for part in normalized(record.get(field)).split(" | ") if part.strip())
        for raw_url in raw_urls:
            url = with_scheme(raw_url)
            if url:
                url_owners.setdefault(url, set()).add(f"ms_staging:{index}:{normalized(record.get('farm_name'))}")
    return {
        "candidate_file": str(MS_CANDIDATES),
        "release_id": payload.get("release_id"),
        "candidate_rows": len(records),
        "county_present": sum(bool(normalized(record.get("county"))) for record in records),
        "city_present": sum(bool(normalized(record.get("city"))) for record in records),
        "source_url_present": sum(bool(normalized(record.get("source_urls"))) for record in records),
        "pending_promotion": sum(normalized(record.get("review_status")).startswith("New candidate") for record in records),
    }, findings


def previous_url_statuses(current_dir: Path) -> dict[str, str]:
    prior_dirs = sorted(
        (path for path in AUDIT_ROOT.iterdir() if path.is_dir() and path.name < current_dir.name),
        key=lambda path: path.name,
    ) if AUDIT_ROOT.exists() else []
    for directory in reversed(prior_dirs):
        report_path = directory / "verification-results.json"
        if report_path.exists():
            payload = json.loads(report_path.read_text())
            return {item["url"]: item["status"] for item in payload.get("url_checks", [])}
    return {}


def main() -> int:
    output_dir = AUDIT_ROOT / TODAY
    output_dir.mkdir(parents=True, exist_ok=True)
    canonical, canonical_findings, url_owners = canonical_audit()
    staging, staging_findings = staging_audit(url_owners)
    previous = previous_url_statuses(output_dir)

    checks: list[dict] = []
    with ThreadPoolExecutor(max_workers=12) as executor:
        futures = {executor.submit(check_url, url): url for url in sorted(url_owners)}
        for future in as_completed(futures):
            result = future.result()
            result["owners"] = " | ".join(sorted(url_owners[result["url"]]))
            result["previous_status"] = previous.get(result["url"], "")
            result["changed_since_prior_audit"] = bool(result["previous_status"] and result["previous_status"] != result["status"])
            checks.append(result)
    checks.sort(key=lambda item: item["url"])

    all_findings = canonical_findings + staging_findings
    summary = {
        "audit_date": TODAY,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "mode": "observe_compare_stage_for_review",
        "retry_policy": "up to 3 attempts per URL; no failed row is silently dropped",
        "canonical": canonical,
        "staging": staging,
        "url_summary": {
            status: sum(item["status"] == status for item in checks)
            for status in sorted({item["status"] for item in checks})
        },
        "url_changes_since_prior_audit": sum(item["changed_since_prior_audit"] for item in checks),
        "row_findings": len(all_findings),
        "promotion_policy": "No automatic canonical changes. Review source evidence and promote an atomic release only after gates pass.",
        "url_checks": checks,
    }
    (output_dir / "verification-results.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    with (output_dir / "url-checks.csv").open("w", newline="", encoding="utf-8") as handle:
        fields = list(checks[0]) if checks else []
        writer = csv.DictWriter(handle, fieldnames=fields)
        if fields:
            writer.writeheader()
            writer.writerows(checks)
    with (output_dir / "row-findings.csv").open("w", newline="", encoding="utf-8") as handle:
        fields = ["dataset", "row", "farm_name", "state", "findings"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_findings)

    concise = {
        "status": "failed" if canonical["errors"] else "passed_with_review_items",
        "audit_date": TODAY,
        "canonical_rows": canonical["source_rows"],
        "staging_rows": staging.get("candidate_rows", 0),
        "urls_checked": len(checks),
        "url_summary": summary["url_summary"],
        "row_findings": len(all_findings),
        "report": str(output_dir / "verification-results.json"),
    }
    print(json.dumps(concise, indent=2))
    return 1 if canonical["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
